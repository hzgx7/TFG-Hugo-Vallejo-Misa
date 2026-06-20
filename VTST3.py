#!/usr/bin/python3
'''
DRPath = distinguish reaction path
*----------------------------------*
| Program    :  drpath             |
| Last Update:  2026/02/19 (Y/M/D) |
| Main Author:  David Ferro-Costas |
| Modificado :  Hugo Vallejo Misa  |
*----------------------------------*
'''
#=====================================================================#
import os
import re
import sys
import glob
import numpy as np
#---------------------------------------------------------------------#
import common.physcons as pc
import common.internal as intl
import common.files    as ff
import common.fncs     as fncs
import common.gaussian as gau
import common.partfns  as pf
from   common.MolGraph  import MolGraph
from   common.Molecule  import Molecule
from   common.Spline    import Spline
#---------------------------------------------------------------------#
import getpass
sys.path.append('/home/programs/Pilgrim-2023.1a/')
from   modpilgrim.fit2anarc import fit2anarc
#=====================================================================#

# ============================================================
# CONFIGURACIÓN — MODIFICA AQUÍ
# ============================================================

SCAN_DIR   = "/mnt/netapp2/Store_uni/home/usc/cq/hmi/C6H6/Scan/MIN44_freq"
fscal      = 1.000
nR         = 1
OUTPUT_CSV = "vtst_scan_min44_pr14_C6H6.xlsx"

Tlist = [100, 200, 300, 400, 500, 600, 700, 800, 900, 1000, 1100, 1200,
         1300, 1400, 1500, 1600, 1700, 1800, 1900, 2000, 2100, 2200,
         2300, 2400, 2500, 3000]

gref = [
    -231.97234962,   # T=100 K
    -231.99616220,   # T=200 K
    -232.02101913,   # T=300 K
    -232.04690045,   # T=400 K
    -232.07381200,   # T=500 K
    -232.10173234,   # T=600 K
    -232.13061969,   # T=700 K
    -232.16042414,   # T=800 K
    -232.19109474,   # T=900 K
    -232.22258264,   # T=1000 K
    -232.25484222,   # T=1100 K
    -232.28783131,   # T=1200 K
    -232.32151113,   # T=1300 K
    -232.35584604,   # T=1400 K
    -232.39080330,   # T=1500 K
    -232.42635280,   # T=1600 K
    -232.46246680,   # T=1700 K
    -232.49911976,   # T=1800 K
    -232.53628805,   # T=1900 K
    -232.57394985,   # T=2000 K
    -232.61208491,   # T=2100 K
    -232.65067446,   # T=2200 K
    -232.68970101,   # T=2300 K
    -232.72914828,   # T=2400 K
    -232.76900109,   # T=2500 K
    -232.97388099,   # T=3000 K
]

SCAN_ATOMS = (4, 9)

# ============================================================
# CORRECCIÓN DE FRECUENCIA MÍNIMA (opcional)
# ─────────────────────────────────────────────────────────────
# FREQ_FIX_FROM_POINT : punto a partir del cual aplicar (None = desactivado)
# FREQ_FIX_VALUE_CM   : valor en cm^-1 para sustituir la frecuencia más baja
#
# Ejemplo: desde el punto 51 en adelante, freq_min = 50 cm^-1
#   FREQ_FIX_FROM_POINT = 51
#   FREQ_FIX_VALUE_CM   = 50.0
#
# Para desactivar:
#   FREQ_FIX_FROM_POINT = None
# ============================================================

FREQ_FIX_FROM_POINT = 51
FREQ_FIX_VALUE_CM   = 50.0

# ============================================================
# SELECCIÓN MANUAL DE PUNTOS (opcional)
# ─────────────────────────────────────────────────────────────
# POINT_RANGE : rango de puntos a procesar [inicio, fin] ambos incluidos.
#               None = procesar todos los que encuentre.
# Ejemplo: solo del 1 al 90:
#   POINT_RANGE = [1, 90]
# Para procesar todos:
#   POINT_RANGE = None
# ─────────────────────────────────────────────────────────────
# POINTS_EXCLUDE : lista de puntos a excluir individualmente.
#                  [] o None = no excluir ninguno.
# Ejemplo: excluir los puntos 15, 32 y 67:
#   POINTS_EXCLUDE = [15, 32, 67]
# Para no excluir ninguno:
#   POINTS_EXCLUDE = []
# ============================================================

POINT_RANGE   = [1, 90]      # ← ej: [1, 90]  o  None
POINTS_EXCLUDE = []       # ← ej: [15, 32, 67]  o  []Esto es que ningun pto

# ============================================================
# FUNCIÓN PRINCIPAL: procesar un punto
# ============================================================

def process_point(log, Tlist, fscal, gref, nR=1, point_num=0):
    if not os.path.exists(log):
        return None, "archivo no existe"
    if not gau.normal_termination(log):
        return None, "terminacion anormal de Gaussian"

    try:
        geom = Molecule()
        geom.set_from_gauout(log)
        geom.setvar(fscal=fscal)
        geom.setup(projgrad=True)
        geom.ana_freqs("cc")

        # ── CORRECCIÓN DE FRECUENCIA MÍNIMA ──────────────────────────
        if (FREQ_FIX_FROM_POINT is not None and
                point_num >= FREQ_FIX_FROM_POINT):

            ccfreqs     = list(geom._ccfreqs)
            min_idx     = min(range(len(ccfreqs)), key=lambda i: ccfreqs[i])
            min_freq_cm = fncs.afreq2cm(ccfreqs[min_idx])
            scale       = FREQ_FIX_VALUE_CM / min_freq_cm

            print("  [FREQ FIX] Punto %d: freq_min original = %.2f cm^-1  →  sustituida por %.2f cm^-1"
                  % (point_num, min_freq_cm, FREQ_FIX_VALUE_CM))

            ccfreqs[min_idx] = ccfreqs[min_idx] * scale
            geom._ccfreqs    = ccfreqs
        # ─────────────────────────────────────────────────────────────

        # Imprimir frecuencias en pantalla
        NNN     = 5
        ccfreqs = list(geom._ccfreqs)
        for ii in range(0, len(ccfreqs)+1, NNN):
            ccfreq = ccfreqs[ii:ii+NNN]
            line   = "".join(["  %7.2f" % (fncs.afreq2cm(f)) for f in ccfreq])
            print(line)
        print("")

        qtot, V1, pfns = geom.calc_pfns(Tlist, "cc")
        gibbscc = V1 - pc.KB * np.array(Tlist) * np.log(np.array(qtot) * pc.VOL0)

        human_units = pc.ML**(nR - 1.0) / pc.SECOND

        results = []
        for idx in range(len(Tlist)):
            T      = Tlist[idx]
            g      = gibbscc[idx]
            gref_i = gref[idx]
            grel   = g - gref_i
            k      = pf.GFEA2rate([T], [grel], nR)[0] * human_units

            print("      * T = %7.2f K : G = %18.8f hartree   [%.3E]" % (T, g, k))

            results.append({
                "T"          : T,
                "G_punto"    : g,
                "G_reactivo" : gref_i,
                "DeltaG"     : grel,
                "k"          : k,
            })
        print("")
        return results, None

    except Exception as e:
        return None, str(e)

# ============================================================
# EXTRAER DISTANCIA DEL SCAN (opcional)
# ============================================================

def extract_distance(log, atom1, atom2):
    if atom1 is None or atom2 is None:
        return None

    a1, a2 = min(atom1, atom2), max(atom1, atom2)
    dist   = None

    with open(log) as f:
        content = f.read()

    blocks = content.split("Distance matrix (angstroms):")
    if len(blocks) < 2:
        return None

    last_block = blocks[-1]
    lines      = last_block.split("\n")

    for line in lines:
        m = re.match(r"\s+%d\s+\w+" % a2, line)
        if m:
            parts   = line.split()
            col_idx = a1 + 1
            if col_idx < len(parts):
                try:
                    dist = float(parts[col_idx])
                except:
                    dist = None
            break

    return dist

# ============================================================
# MAIN
# ============================================================

def main():

    pattern = os.path.join(SCAN_DIR, "freq_*.out")
    files   = glob.glob(pattern)

    def get_num(f):
        m = re.search(r"freq_(\d+)\.out", os.path.basename(f))
        return int(m.group(1)) if m else 0

    files = sorted(files, key=get_num)

    if not files:
        print("No se encontraron archivos freq_*.out en:", SCAN_DIR)
        return

    # ── APLICAR FILTROS DE SELECCIÓN MANUAL ──────────────────
    exclude_set = set(POINTS_EXCLUDE) if POINTS_EXCLUDE else set()

    filtered = []
    skipped  = []
    for f in files:
        num = get_num(f)

        # Filtro por rango
        if POINT_RANGE is not None:
            if num < POINT_RANGE[0] or num > POINT_RANGE[1]:
                skipped.append(num)
                continue

        # Filtro por exclusión manual
        if num in exclude_set:
            skipped.append(num)
            continue

        filtered.append(f)

    files = filtered
    # ─────────────────────────────────────────────────────────

    print("=" * 60)
    print("  DRPath - procesando %d archivos" % len(files))
    if POINT_RANGE is not None:
        print("  [RANGO] Del punto %d al %d" % (POINT_RANGE[0], POINT_RANGE[1]))
    if exclude_set:
        print("  [EXCLUIDOS] Puntos: %s" % sorted(exclude_set))
    if skipped:
        print("  [SALTADOS] %d puntos fuera de rango o excluidos" % len(skipped))
    if FREQ_FIX_FROM_POINT is not None:
        print("  [FREQ FIX] Desde punto %d → freq_min = %.1f cm^-1"
              % (FREQ_FIX_FROM_POINT, FREQ_FIX_VALUE_CM))
    print("=" * 60)

    csv_rows    = []
    ok_count    = 0
    fail_count  = 0
    failed_list = []

    for log in files:
        num = get_num(log)
        print("\n" + "=" * 60)
        print("  Punto %d : %s" % (num, os.path.basename(log)))
        print("=" * 60)

        dist = None
        if SCAN_ATOMS is not None:
            dist = extract_distance(log, SCAN_ATOMS[0], SCAN_ATOMS[1])

        results, error = process_point(log, Tlist, fscal, gref, nR,
                                       point_num=num)

        if results is None:
            print("  !! FALLO: %s" % error)
            fail_count += 1
            failed_list.append((num, os.path.basename(log), error))
            continue

        ok_count += 1
        for r in results:
            csv_rows.append({
                "T"          : r["T"],
                "Punto"      : num,
                "Distancia"  : dist,
                "G_reactivo" : r["G_reactivo"],
                "G_punto"    : r["G_punto"],
                "DeltaG"     : r["DeltaG"],
                "k"          : r["k"],
            })

    print("\n" + "=" * 60)
    print("  RESUMEN FINAL")
    print("=" * 60)
    print("  Archivos procesados OK : %d" % ok_count)
    print("  Archivos con fallo     : %d" % fail_count)
    if failed_list:
        print("\n  Archivos fallidos:")
        for num, fname, err in failed_list:
            print("    - Punto %d (%s): %s" % (num, fname, err))
    print("=" * 60)

    if not csv_rows:
        print("\nNo hay datos para escribir.")
        return

    from collections import defaultdict
    from openpyxl import Workbook
    from openpyxl.chart import ScatterChart, Reference, Series
    from openpyxl.styles import Font

    by_temp = defaultdict(list)
    for r in csv_rows:
        by_temp[r["T"]].append(r)
    for T in by_temp:
        by_temp[T] = sorted(by_temp[T], key=lambda x: x["Punto"])

    print("\n=== MAXIMO DeltaG POR TEMPERATURA ===")
    for T in sorted(by_temp.keys()):
        rows    = by_temp[T]
        max_row = max(rows, key=lambda x: x["DeltaG"])
        print("  T = %7.2f K : Punto = %d  Distancia = %.4f A  DeltaG = %+.8f hartree  k = %.3E s-1" % (
            T, max_row["Punto"],
            max_row["Distancia"] if max_row["Distancia"] else 0,
            max_row["DeltaG"], max_row["k"]))

    wb = Workbook()
    ws = wb.active
    ws.title = "VTST_scan"
    current_row = 1

    for T in sorted(by_temp.keys()):
        rows = by_temp[T]

        ws.cell(row=current_row, column=1,
                value="T = %.2f K" % T).font = Font(bold=True)
        current_row += 1

        headers = ["Punto", "Distancia (A)", "G_reactivo (hartree)",
                   "G_punto (hartree)", "DeltaG (hartree)", "k (s-1)"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=current_row, column=col, value=h).font = Font(bold=True)
        header_row  = current_row
        current_row += 1

        data_start = current_row
        for r in rows:
            d = r["Distancia"] if r["Distancia"] is not None else 0.0
            ws.cell(row=current_row, column=1, value=r["Punto"])
            ws.cell(row=current_row, column=2, value=round(d,               6))
            ws.cell(row=current_row, column=3, value=round(r["G_reactivo"], 8))
            ws.cell(row=current_row, column=4, value=round(r["G_punto"],    8))
            ws.cell(row=current_row, column=5, value=round(r["DeltaG"],     8))
            ws.cell(row=current_row, column=6, value=r["k"])
            current_row += 1
        data_end = current_row - 1

        chart = ScatterChart()
        chart.title        = "DeltaG vs Distancia  —  T = %g K" % T
        chart.style        = 10
        chart.x_axis.title = "Distancia (Angstrom)"
        chart.y_axis.title = "DeltaG (hartree)"
        chart.legend       = None

        xvalues = Reference(ws, min_col=2, min_row=data_start, max_row=data_end)
        yvalues = Reference(ws, min_col=5, min_row=data_start, max_row=data_end)

        series = Series(yvalues, xvalues, title="DeltaG")
        series.marker.symbol = "circle"
        series.marker.size   = 4
        series.graphicalProperties.line.solidFill = "1F77B4"
        chart.series.append(series)

        chart.width  = 20
        chart.height = 14
        ws.add_chart(chart, "H%d" % header_row)

        current_row += 2

    OUTPUT_XLSX = OUTPUT_CSV.replace(".csv", ".xlsx")
    wb.save(OUTPUT_XLSX)
    print("\nArchivo guardado: %s" % OUTPUT_XLSX)
    print("Proceso completado correctamente ✔\n")

if __name__ == "__main__":
    main()
